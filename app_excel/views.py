import io
import json
import os
import re
import zipfile
from datetime import datetime
from urllib.parse import quote_plus

import openpyxl
from docx import Document
from docxtpl import DocxTemplate
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from rest_framework import viewsets

from .models import Profile, Student
from .serializers import StudentSerializer


UZBEK_MONTHS = [
    "yanvar",
    "fevral",
    "mart",
    "aprel",
    "may",
    "iyun",
    "iyul",
    "avgust",
    "sentabr",
    "oktabr",
    "noyabr",
    "dekabr",
]

REQUIRED_STUDENT_FIELDS = [
    ("full_name", "Talabaning F.I.Sh."),
    ("faculty", "Fakultet"),
    ("direction", "Yo'nalish"),
    ("group", "Guruh"),
    ("company", "Korxona nomi"),
    ("company_address", "Korxona manzili"),
    ("company_director", "Korxona rahbari F.I.Sh."),
    ("company_phone", "Korxona telefoni"),
    ("practice_supervisor", "Universitet amaliyot rahbari"),
    ("faculty_dean", "Fakultet dekani"),
    ("department_head", "Kafedra mudiri"),
]

DOCUMENT_ARCHIVE_TYPES = [
    {"key": "shartnoma", "label": "Shartnoma"},
    {"key": "kundalik", "label": "Kundalik"},
    {"key": "yollanma", "label": "Yo'llanma"},
]

STUDENT_IMPORT_UPDATE_FIELDS = [
    "full_name",
    "direction",
    "faculty",
    "group",
    "company",
    "company_address",
    "company_director",
    "company_phone",
    "practice_supervisor",
    "faculty_dean",
    "department_head",
]

DEFAULT_CHAT_RESPONSES = [
    (
        ("salom", "assalom", "hello", "hi"),
        "Assalomu alaykum! Men e-Praktika yordamchisiman. Excel yuklash, hujjatlar, talabalar ro'yxati, korxonalar arxivi yoki profil sozlamalari bo'yicha yordam beraman.",
    ),
    (
        ("excel", "yuklash", "upload", "fayl"),
        "Excel fayl yuklash tartibi:\n1. /excel/upload/ sahifasiga o'ting.\n2. Namuna Excel faylini yuklab oling.\n3. Talabalar ma'lumotlarini shu ustunlar bo'yicha to'ldiring.\n4. Faylni .xlsx formatda yuklang.\n\nYangi talabalar mavjud ro'yxatga qo'shiladi. Oldingi talabalar o'chib ketmaydi.",
    ),
    (
        ("namuna", "shablon", "sample"),
        "Namuna Excel kerak bo'lsa, /excel/upload/ sahifasidagi Namuna Excel tugmasini bosing. Fayl ichida kerakli ustunlar va to'ldirish namunasi bor.",
    ),
    (
        ("hujjat", "shartnoma", "kundalik", "yollanma", "zip", "word", "docx"),
        "Hujjat olish tartibi:\n1. Avval Excel ro'yxatini yuklang.\n2. /excel/documents/ sahifasiga o'ting.\n3. Talaba bo'yicha shartnoma, kundalik yoki yo'llanmani yuklab oling.\n4. Kerak bo'lsa barcha hujjatlarni ZIP qilib oling.",
    ),
    (
        ("talaba", "student", "royxat", "ro'yxat"),
        "Talabalar ro'yxati /excel/students/ sahifasida ko'rinadi. Yangi Excel yuklanganda talabalar mavjud ro'yxatga qo'shiladi, oldingi ma'lumotlar saqlanadi.",
    ),
    (
        ("korxona", "kompaniya", "company", "tashkilot"),
        "Korxonalar arxivi /excel/companies/ sahifasida. Bu yerda korxona nomi, manzili, telefoni va biriktirilgan talabalar ko'rinadi.",
    ),
    (
        ("profil", "akkaunt", "sozlama", "parol", "password"),
        "Profil ma'lumotlari uchun /excel/profile/ sahifasiga o'ting. Parol yoki akkaunt sozlamalarini o'zgartirish uchun /excel/settings/ sahifasidan foydalaning.",
    ),
    (
        ("login", "kirish", "register", "royxatdan", "ro'yxatdan"),
        "Tizimga kirish uchun /excel/login/ sahifasidan foydalaning. Yangi akkaunt ochish uchun /excel/register/ sahifasiga o'ting.",
    ),
    (
        ("admin", "operator"),
        "Admin panel /admin/ manzilida. Default foydalanuvchilar yaratilgan bo'lsa, admin yoki operator akkaunti orqali kirish mumkin.",
    ),
    (
        ("xato", "error", "ishlamayapti", "muammo"),
        "Agar xatolik chiqsa:\n1. Sahifani yangilang.\n2. Excel fayl .xlsx formatda ekanini tekshiring.\n3. Majburiy ustunlar to'ldirilganini tekshiring.\n4. Muammo davom etsa, quyidagi kontaktlar orqali murojaat qiling.",
    ),
]

DEFAULT_CHAT_FALLBACK = """
Savolingizni to'liq tushunmadim. Quyidagilardan birini yozib ko'ring:

- Excel qanday yuklanadi?
- Hujjatlarni qayerdan olaman?
- Talabalar ro'yxati qayerda?
- Korxonalar arxivi qayerda?
- Profil yoki parol qanday o'zgartiriladi?
""".strip()


def clean_chat_text(value, max_length=1200):
    return " ".join(str(value or "").strip().split())[:max_length]


def format_support_phone(phone):
    digits = "".join(char for char in phone if char.isdigit())
    if len(digits) == 9:
        digits = f"998{digits}"

    if len(digits) == 12 and digits.startswith("998"):
        return f"+998 {digits[3:5]} {digits[5:8]} {digits[8:10]} {digits[10:12]}"

    return phone.strip()


def get_support_contact_note():
    contacts = []
    telegram = getattr(settings, "SUPPORT_TELEGRAM", "").strip()
    if telegram:
        if telegram.startswith("@"):
            telegram = f"{telegram} (https://t.me/{telegram[1:]})"
        contacts.append(f"Telegram: {telegram}")

    phone = format_support_phone(getattr(settings, "SUPPORT_PHONE", "").strip())
    if phone:
        contacts.append(f"Telefon: {phone}")

    if contacts:
        return "Qo'shimcha savol yoki tushunmovchilik bo'lsa, murojaat qiling:\n" + "\n".join(contacts)

    return "Qo'shimcha savol yoki tushunmovchilik bo'lsa, administratorga Telegram orqali murojaat qiling."


def add_support_note(reply):
    return f"{reply}\n\n{get_support_contact_note()}"


def message_has_keyword(message, keyword):
    if len(keyword) <= 3:
        return re.search(rf"(^|[^a-z0-9_]){re.escape(keyword)}([^a-z0-9_]|$)", message) is not None
    return keyword in message


def get_default_chat_reply(message):
    normalized_message = message.lower()
    for keywords, reply in DEFAULT_CHAT_RESPONSES:
        if any(message_has_keyword(normalized_message, keyword) for keyword in keywords):
            if any(message_has_keyword(normalized_message, keyword) for keyword in ("xato", "error", "ishlamayapti", "muammo")):
                return add_support_note(reply)
            return reply
    return add_support_note(DEFAULT_CHAT_FALLBACK)


def ensure_profile(user):
    return Profile.objects.get_or_create(user=user)[0]


def get_template_sources():
    source_dir = os.path.join(settings.BASE_DIR, "app_excel", "document_sources")
    return [
        {
            "key": "shartnoma",
            "label": "Shartnoma",
            "filename": "Bitiruv oldi Shartnomasi.doc",
            "kind": "DOC manba",
            "exists": os.path.exists(os.path.join(source_dir, "Bitiruv oldi Shartnomasi.doc")),
        },
        {
            "key": "kundalik",
            "label": "Kundalik",
            "filename": "Bitiruv oldi amaliyot Kundalik.doc",
            "kind": "DOC manba",
            "exists": os.path.exists(os.path.join(source_dir, "Bitiruv oldi amaliyot Kundalik.doc")),
        },
        {
            "key": "yollanma",
            "label": "Yo'llanma",
            "filename": "Bitiruv oldi Yo'llanma.doc",
            "kind": "DOC manba",
            "exists": os.path.exists(os.path.join(source_dir, "Bitiruv oldi Yo'llanma.doc")),
        },
    ]


def get_runtime_template_paths():
    template_dir = os.path.join(settings.BASE_DIR, "app_excel", "templates", "app_excel")
    return {
        "shartnoma": os.path.join(template_dir, "shartnoma_template.docx"),
        "kundalik": os.path.join(template_dir, "kundalik_template.docx"),
        "yollanma": os.path.join(template_dir, "yollanma_template.docx"),
    }


def get_available_runtime_template_path(template_key):
    path = get_runtime_template_paths()[template_key]
    return path if os.path.exists(path) else None


def get_template_variables(template_key):
    template_path = get_available_runtime_template_path(template_key)
    if not template_path:
        return set()
    template = DocxTemplate(template_path)
    return template.get_undeclared_template_variables()


def validate_runtime_templates():
    issues = []
    for template_key, template_path in get_runtime_template_paths().items():
        if not os.path.exists(template_path):
            issues.append(f"{template_key} topilmadi")
            continue
        if not get_template_variables(template_key):
            issues.append(f"{template_key} ichida {{ }} maydonlari yo'q")
    return issues


def get_practice_type(course):
    return "Ishlab chiqarish amaliyoti" if int(course) == 3 else "Bitiruv oldi amaliyoti"


def get_company_student_count(company_name):
    return Student.objects.filter(company=company_name).count()


def build_company_archive(students):
    companies = {}
    for student in students:
        company_name = (student.company or "").strip() or "Nomi kiritilmagan"
        company = companies.setdefault(
            company_name,
            {
                "name": company_name,
                "student_count": 0,
                "director": "",
                "address": "",
                "phone": "",
                "students": [],
            },
        )
        company["student_count"] += 1
        company["students"].append(student.full_name)
        company["director"] = company["director"] or student.company_director
        company["address"] = company["address"] or student.company_address
        company["phone"] = company["phone"] or student.company_phone

    for company in companies.values():
        map_query = f"{company['name']} {company['address']}".strip()
        company["map_query"] = quote_plus(map_query or company["name"])
    return sorted(companies.values(), key=lambda company: company["name"].lower())


def get_missing_student_fields(student):
    missing = []
    for field_name, label in REQUIRED_STUDENT_FIELDS:
        if not str(getattr(student, field_name, "") or "").strip():
            missing.append(label)
    return missing


def normalize_student_identity_value(value):
    return " ".join(str(value or "").strip().lower().split())


def get_student_import_key(student):
    return (
        normalize_student_identity_value(student.full_name),
        normalize_student_identity_value(student.group),
        normalize_student_identity_value(student.company),
    )


def copy_student_import_fields(target, source):
    changed = False
    for field_name in STUDENT_IMPORT_UPDATE_FIELDS:
        source_value = getattr(source, field_name)
        if getattr(target, field_name) != source_value:
            setattr(target, field_name, source_value)
            changed = True
    return changed


def upsert_imported_students(students):
    existing_by_key = {}
    for student in Student.objects.all().order_by("id"):
        key = get_student_import_key(student)
        if key not in existing_by_key:
            existing_by_key[key] = student

    incoming_by_key = {}
    for student in students:
        incoming_by_key[get_student_import_key(student)] = student

    students_to_create = []
    updated_count = 0
    for key, student in incoming_by_key.items():
        existing_student = existing_by_key.get(key)
        if existing_student is None:
            students_to_create.append(student)
            continue

        changed = copy_student_import_fields(existing_student, student)
        if changed:
            existing_student.save(update_fields=STUDENT_IMPORT_UPDATE_FIELDS)
        updated_count += 1

    if students_to_create:
        Student.objects.bulk_create(students_to_create)

    return {
        "added": len(students_to_create),
        "updated": updated_count,
        "skipped_duplicates": len(students) - len(incoming_by_key),
    }


def validate_students_for_documents(students, course, start_date_str, end_date_str):
    issues = []
    if not course:
        issues.append("Kurs tanlanmagan.")
    if not start_date_str or not end_date_str:
        issues.append("Boshlanish va tugash sanalari to'liq kiritilmagan.")

    for index, student in enumerate(students, start=1):
        missing = get_missing_student_fields(student)
        if missing:
            issues.append(f"{index}. {student.full_name}: {', '.join(missing)}")
    return issues


def sanitize_filename(value):
    safe = "".join(char if char.isalnum() or char in (" ", "-", "_") else "_" for char in (value or "").strip())
    return "_".join(safe.split()) or "talaba"


def normalize_excel_header(value):
    return " ".join(str(value or "").strip().lower().split())


def get_excel_value(row_map, *aliases):
    for alias in aliases:
        normalized_alias = normalize_excel_header(alias)
        if normalized_alias in row_map:
            return str(row_map.get(normalized_alias) or "").strip()
    return ""


def normalize_excel_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass

    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def parse_date_parts(date_str):
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    return {
        "day": f"{date_obj.day:02}",
        "day_int": str(date_obj.day),
        "month": UZBEK_MONTHS[date_obj.month - 1],
        "year": str(date_obj.year),
        "short": date_obj.strftime("%d.%m.%Y"),
    }


def format_document_date(date_str):
    parts = parse_date_parts(date_str)
    return f"{parts['year']}-yil {parts['day_int']}-{parts['month']}"


def build_common_template_context(student, course, practice_type, start_date_str, end_date_str):
    start_parts = parse_date_parts(start_date_str)
    end_parts = parse_date_parts(end_date_str)
    direction_label = student.direction or student.faculty
    company_student_count = get_company_student_count(student.company)
    academic_year = f"{start_parts['year']}/{int(start_parts['year']) + 1}"
    practice_duration_rule = (
        "Amaliyot muddati 3-bosqichda o'quv reja asosida belgilanadi."
        if int(course) == 3
        else "Amaliyot muddati 4-bosqichda 10-hafta, haftasiga 3 ish kuni, kuniga 6 soat, jami 180 soat etib belgilanadi."
    )

    context = {
        "FULL_NAME": student.full_name,
        "DIRECTION": direction_label,
        "FACULTY": student.faculty,
        "COURSE": str(course),
        "PRACTICE_TYPE": practice_type,
        "PRACTICE_TYPE_LOWER": practice_type.lower(),
        "PRACTICE_TYPE_UPPER": practice_type.upper(),
        "PRACTICE_DURATION_RULE": practice_duration_rule,
        "COMPANY": student.company,
        "ADDRESS": student.company_address,
        "DIRECTOR": student.company_director,
        "PHONE": student.company_phone,
        "SUPERVISOR": student.practice_supervisor,
        "FACULTY_DEAN": student.faculty_dean,
        "DEPARTMENT_HEAD": student.department_head,
        "COMPANY_STUDENT_COUNT": str(company_student_count),
        "ACADEMIC_YEAR": academic_year,
        "START_DAY": start_parts["day"],
        "START_MONTH": start_parts["month"],
        "START_YEAR": start_parts["year"],
        "START_DATE_SHORT": start_parts["short"],
        "START_DATE_TEXT": format_document_date(start_date_str),
        "END_DAY": end_parts["day"],
        "END_MONTH": end_parts["month"],
        "END_YEAR": end_parts["year"],
        "END_DATE_SHORT": end_parts["short"],
        "END_DATE_TEXT": format_document_date(end_date_str),
        "POSITION": "Amaliyotchi",
        "GRADE": "5",
    }
    return context


def render_docx_template(template_key, context):
    template_path = get_available_runtime_template_path(template_key)
    if not template_path:
        raise FileNotFoundError(f"{template_key} shabloni topilmadi")

    template = DocxTemplate(template_path)
    if not template.get_undeclared_template_variables():
        raise ValueError(f"{template_key} shablonida to'ldiriladigan maydonlar yo'q")

    template.render(context)
    buffer = io.BytesIO()
    template.save(buffer)
    buffer.seek(0)
    return Document(buffer)


def generate_document_from_template(template_key, student, course, practice_type, start_date_str, end_date_str):
    context = build_common_template_context(student, course, practice_type, start_date_str, end_date_str)
    return render_docx_template(template_key, context)


def generate_shartnoma_document(student, course, practice_type, start_date_str, end_date_str):
    return generate_document_from_template("shartnoma", student, course, practice_type, start_date_str, end_date_str)


def generate_kundalik_document(student, course, practice_type, start_date_str, end_date_str):
    return generate_document_from_template("kundalik", student, course, practice_type, start_date_str, end_date_str)


def generate_yollanma_document(student, course, practice_type, start_date_str, end_date_str):
    return generate_document_from_template("yollanma", student, course, practice_type, start_date_str, end_date_str)


def save_document_response(document, filename):
    buffer = io.BytesIO()
    document.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_documents_zip(students, course, practice_type, start_date_str, end_date_str):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for index, student in enumerate(students, start=1):
            folder = f"Hujjatlar/{index:02}_{sanitize_filename(student.full_name)}"
            generated_docs = [
                ("shartnoma", generate_shartnoma_document(student, course, practice_type, start_date_str, end_date_str)),
                ("kundalik", generate_kundalik_document(student, course, practice_type, start_date_str, end_date_str)),
                ("yollanma", generate_yollanma_document(student, course, practice_type, start_date_str, end_date_str)),
            ]
            for filename, generated_doc in generated_docs:
                temp_buffer = io.BytesIO()
                generated_doc.save(temp_buffer)
                temp_buffer.seek(0)
                zip_file.writestr(f"{folder}/{filename}.docx", temp_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer


@login_required(login_url="login")
def dashboard_view(request):
    profile = ensure_profile(request.user)
    students = Student.objects.all()
    student_page = Paginator(students.order_by("-id"), 3).get_page(request.GET.get("students_page"))
    return render(
        request,
        "app_excel/dashboard.html",
        {
            "profile": profile,
            "student_count": students.count(),
            "enterprise_count": students.values("company").distinct().count(),
            "document_count": students.count() * len(DOCUMENT_ARCHIVE_TYPES),
            "recent_students": student_page.object_list,
            "student_page": student_page,
        },
    )


@login_required(login_url="login")
def students_archive_view(request):
    profile = ensure_profile(request.user)
    students = Student.objects.all()
    student_page = Paginator(students.order_by("-id"), 3).get_page(request.GET.get("students_page"))
    return render(
        request,
        "app_excel/students_archive.html",
        {
            "profile": profile,
            "recent_students": student_page.object_list,
            "student_page": student_page,
            "student_count": students.count(),
        },
    )


@login_required(login_url="login")
def companies_archive_view(request):
    profile = ensure_profile(request.user)
    students = list(Student.objects.all().order_by("company", "full_name"))
    company_archive = build_company_archive(students)
    return render(
        request,
        "app_excel/companies_archive.html",
        {
            "profile": profile,
            "company_archive": company_archive,
            "enterprise_count": len(company_archive),
        },
    )


@login_required(login_url="login")
def documents_archive_view(request):
    profile = ensure_profile(request.user)
    students = Student.objects.all().order_by("full_name")
    return render(
        request,
        "app_excel/documents_archive.html",
        {
            "profile": profile,
            "students_archive": students,
            "document_types": DOCUMENT_ARCHIVE_TYPES,
            "document_count": students.count() * len(DOCUMENT_ARCHIVE_TYPES),
        },
    )


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer


@login_required(login_url="login")
@require_POST
def ai_chat_view(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Noto'g'ri JSON yuborildi."}, status=400)

    message = clean_chat_text(payload.get("message"), 1000)

    if not message:
        return JsonResponse({"error": "Savol matnini kiriting."}, status=400)

    return JsonResponse({"reply": get_default_chat_reply(message)})


@login_required(login_url="login")
def download_sample_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Talabalar"

    headers = [
        "Talabaning F.I.Sh.",
        "Fakultet",
        "Yo'nalish",
        "Guruh",
        "Kurs",
        "Amaliyot turi",
        "Boshlanish sanasi",
        "Tugash sanasi",
        "Korxona nomi",
        "Korxona manzili",
        "Korxona rahbari F.I.Sh.",
        "Korxona telefoni",
        "Universitet amaliyot rahbari",
        "Fakultet dekani",
        "Kafedra mudiri",
    ]
    sample_rows = [
        [
            "Aliyev Bekzod Anvar o'g'li",
            "Axborot texnologiyalari",
            "Dasturiy injiniring",
            "DI-401",
            4,
            "Bitiruv oldi amaliyoti",
            "2025-02-17",
            "2025-04-26",
            "TechSoft MCHJ",
            "Toshkent sh., Yunusobod tumani, Amir Temur ko'chasi 24-uy",
            "Karimov Sardor Rustamovich",
            "+998901112233",
            "Rasulov Dilshod Qodirovich",
            "O.B. Ro'zibayev",
            "N.O. Raximov",
        ],
        [
            "Karimova Maftuna Jamshid qizi",
            "Axborot texnologiyalari",
            "Kompyuter injiniring",
            "KI-302",
            3,
            "Ishlab chiqarish amaliyoti",
            "2025-06-02",
            "2025-07-12",
            "Digital Systems MCHJ",
            "Toshkent sh., Chilonzor tumani, Bunyodkor ko'chasi 18-uy",
            "Ergashev Oybek Bahodirovich",
            "+998901112234",
            "Rasulov Dilshod Qodirovich",
            "O.B. Ro'zibayev",
            "N.O. Raximov",
        ],
        [
            "Tursunov Azizbek Ilhom o'g'li",
            "Iqtisodiyot",
            "Raqamli iqtisodiyot",
            "RI-403",
            4,
            "Bitiruv oldi amaliyoti",
            "2025-02-17",
            "2025-04-26",
            "Agrobank ATB",
            "Toshkent sh., Shayxontohur tumani, Navoiy ko'chasi 30-uy",
            "To'xtayev Jamshid Abduvaliyevich",
            "+998901112235",
            "Sattorov Akmal Nabiyevich",
            "D.M. Abdullayeva",
            "S.R. Ismoilov",
        ],
        [
            "Saidova Mohinur Baxtiyor qizi",
            "Iqtisodiyot",
            "Moliya va moliyaviy texnologiyalar",
            "MT-301",
            3,
            "Ishlab chiqarish amaliyoti",
            "2025-06-02",
            "2025-07-12",
            "Hamkorbank ATB",
            "Andijon sh., Bobur shoh ko'chasi 52-uy",
            "Qodirov Lazizbek Muzaffarovich",
            "+998901112236",
            "Sattorov Akmal Nabiyevich",
            "D.M. Abdullayeva",
            "S.R. Ismoilov",
        ],
        [
            "Nazarov Shoxrux Zafar o'g'li",
            "Telekommunikatsiya texnologiyalari",
            "Telekommunikatsiya injiniringi",
            "TI-404",
            4,
            "Bitiruv oldi amaliyoti",
            "2025-02-17",
            "2025-04-26",
            "Uztelecom AK",
            "Toshkent sh., Mirzo Ulug'bek tumani, Mustaqillik shoh ko'chasi 28-uy",
            "Ahmedov Farrux Xamidovich",
            "+998901112237",
            "Yusupov Javlon Shavkatovich",
            "A.T. Mahmudov",
            "B.K. Normatov",
        ],
        [
            "Qodirova Dilnoza Sherali qizi",
            "Telekommunikatsiya texnologiyalari",
            "Axborot xavfsizligi",
            "AX-303",
            3,
            "Ishlab chiqarish amaliyoti",
            "2025-06-02",
            "2025-07-12",
            "CyberSec Solutions MCHJ",
            "Samarqand sh., Rudakiy ko'chasi 77-uy",
            "Nishonov Abror G'ayratovich",
            "+998901112238",
            "Yusupov Javlon Shavkatovich",
            "A.T. Mahmudov",
            "B.K. Normatov",
        ],
        [
            "Rustamov Diyorbek Bahrom o'g'li",
            "Energetika",
            "Elektr energetikasi",
            "EE-402",
            4,
            "Bitiruv oldi amaliyoti",
            "2025-02-17",
            "2025-04-26",
            "Hududiy elektr tarmoqlari AJ",
            "Farg'ona sh., Al-Farg'oniy ko'chasi 14-uy",
            "Mirzayev Otabek Raufovich",
            "+998901112239",
            "Xolmatov Ulug'bek Sobirovich",
            "M.I. Xudoyberdiyev",
            "F.A. Umarov",
        ],
        [
            "Ismoilova Sevinch Ravshan qizi",
            "Energetika",
            "Muqobil energiya manbalari",
            "ME-304",
            3,
            "Ishlab chiqarish amaliyoti",
            "2025-06-02",
            "2025-07-12",
            "Solar Energy Group MCHJ",
            "Namangan sh., Do'stlik shoh ko'chasi 9-uy",
            "Abdurahmonov Shohjahon Ikromovich",
            "+998901112240",
            "Xolmatov Ulug'bek Sobirovich",
            "M.I. Xudoyberdiyev",
            "F.A. Umarov",
        ],
        [
            "Mamatqulov Jasurbek Olim o'g'li",
            "Transport tizimlari",
            "Logistika",
            "LG-405",
            4,
            "Bitiruv oldi amaliyoti",
            "2025-02-17",
            "2025-04-26",
            "UzAuto Motors AJ",
            "Asaka sh., Bobur ko'chasi 73-uy",
            "Xaydarov Bekzod G'ulomovich",
            "+998901112241",
            "Norboyev Sanjar Sobitovich",
            "R.X. Jo'rayev",
            "I.S. To'raqulov",
        ],
        [
            "Xudoyberdiyeva Madina Ulug'bek qizi",
            "Transport tizimlari",
            "Avtomobil servisi",
            "AS-305",
            3,
            "Ishlab chiqarish amaliyoti",
            "2025-06-02",
            "2025-07-12",
            "Express Logistics MCHJ",
            "Buxoro sh., Ibn Sino ko'chasi 41-uy",
            "Sobirov Davron Akmalovich",
            "+998901112242",
            "Norboyev Sanjar Sobitovich",
            "R.X. Jo'rayev",
            "I.S. To'raqulov",
        ],
    ]

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="4338CA")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = openpyxl.styles.Side(style="thin", color="000000")
    medium_side = openpyxl.styles.Side(style="medium", color="000000")
    data_alignment = openpyxl.styles.Alignment(vertical="center", horizontal="left")
    data_fill = openpyxl.styles.PatternFill("solid", fgColor="F8FAFC")

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = openpyxl.styles.Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = max(len(header) + 4, 20)

    for row_index, row_data in enumerate(sample_rows, start=2):
        for column_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = data_alignment
            cell.fill = data_fill
            cell.border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        sheet.row_dimensions[row_index].height = 22

    guide_sheet = workbook.create_sheet("Yoriqnoma")
    guide_rows = [
        ["Ustun", "Izoh", "Majburiy"],
        ["Talabaning F.I.Sh.", "Talabaning to'liq F.I.Sh.", "Ha"],
        ["Fakultet", "Masalan: Axborot texnologiyalari", "Ha"],
        ["Yo'nalish", "Masalan: Dasturiy injiniring", "Ha"],
        ["Guruh", "Masalan: SE-401", "Ha"],
        ["Kurs", "3 yoki 4", "Ha"],
        ["Amaliyot turi", "Masalan: Bitiruv oldi amaliyoti", "Ha"],
        ["Boshlanish sanasi", "YYYY-MM-DD formatda", "Ha"],
        ["Tugash sanasi", "YYYY-MM-DD formatda", "Ha"],
        ["Korxona nomi", "Hujjatlarda chiqadigan tashkilot nomi", "Ha"],
        ["Korxona manzili", "To'liq manzil", "Ha"],
        ["Korxona rahbari F.I.Sh.", "Korxona rahbari yoki korxonadagi amaliyot rahbari", "Ha"],
        ["Korxona telefoni", "Masalan: +998901112233", "Ha"],
        ["Universitet amaliyot rahbari", "Universitetdagi amaliyot rahbari", "Ha"],
        ["Fakultet dekani", "Yo'llanma uchun kerak", "Ha"],
        ["Kafedra mudiri", "Kundalik va yo'llanma uchun kerak", "Ha"],
    ]
    for row_index, row_data in enumerate(guide_rows, start=1):
        for column_index, value in enumerate(row_data, start=1):
            cell = guide_sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = openpyxl.styles.Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            if row_index == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.fill = data_fill

    guide_sheet.column_dimensions["A"].width = 28
    guide_sheet.column_dimensions["B"].width = 48
    guide_sheet.column_dimensions["C"].width = 14
    guide_sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(sample_rows) + 1}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="namuna_talabalar.xlsx"'
    return response


@login_required(login_url="login")
def download_template_source(request, filename):
    allowed_files = {item["key"]: item["filename"] for item in get_template_sources()}
    target_name = allowed_files.get(filename)
    if not target_name:
        return HttpResponse("Shablon topilmadi.", status=404)

    file_path = os.path.join(settings.BASE_DIR, "app_excel", "document_sources", target_name)
    if not os.path.exists(file_path):
        return HttpResponse("Shablon fayli topilmadi.", status=404)

    return FileResponse(open(file_path, "rb"), as_attachment=True, filename=target_name)


@login_required(login_url="login")
def upload_excel(request):
    profile = ensure_profile(request.user)
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]
        filename = uploaded_file.name.lower()
        course = None
        start_date = ""
        end_date = ""
        practice_type = ""

        try:
            if not filename.endswith(".xlsx"):
                raise ValueError("Faqat .xlsx Excel fayl yuklash mumkin.")

            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                raise ValueError("Excel faylda sarlavha qatori topilmadi.")

            normalized_headers = [normalize_excel_header(value) for value in header_row]
            new_students = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_map = {
                    normalized_headers[index]: row[index]
                    for index in range(min(len(normalized_headers), len(row)))
                    if normalized_headers[index]
                }
                full_name = get_excel_value(row_map, "Talabaning F.I.Sh.", "F.I.Sh.", "Talaba F.I.Sh.")
                if not full_name:
                    continue

                if course is None:
                    raw_course = get_excel_value(row_map, "Kurs")
                    course = int(raw_course) if raw_course else 4
                if not start_date:
                    start_date = normalize_excel_date(row_map.get(normalize_excel_header("Boshlanish sanasi")))
                if not end_date:
                    end_date = normalize_excel_date(row_map.get(normalize_excel_header("Tugash sanasi")))
                if not practice_type:
                    practice_type = get_excel_value(row_map, "Amaliyot turi")

                new_students.append(
                    Student(
                        full_name=full_name,
                        direction=get_excel_value(row_map, "Yo'nalish"),
                        faculty=get_excel_value(row_map, "Fakultet"),
                        group=get_excel_value(row_map, "Guruh"),
                        company=get_excel_value(row_map, "Korxona nomi"),
                        company_address=get_excel_value(row_map, "Korxona manzili"),
                        company_director=get_excel_value(row_map, "Korxona rahbari F.I.Sh."),
                        company_phone=get_excel_value(row_map, "Korxona telefoni"),
                        practice_supervisor=get_excel_value(row_map, "Universitet amaliyot rahbari"),
                        faculty_dean=get_excel_value(row_map, "Fakultet dekani"),
                        department_head=get_excel_value(row_map, "Kafedra mudiri"),
                    )
                )

            if not new_students:
                raise ValueError("Excel faylda kamida bitta talaba qatori bo'lishi kerak.")
            if not start_date or not end_date:
                raise ValueError("Excel faylda 'Boshlanish sanasi' va 'Tugash sanasi' ustunlarini to'ldiring.")
            if not practice_type:
                practice_type = get_practice_type(course)

            validation_issues = validate_students_for_documents(new_students, course, start_date, end_date)
            if validation_issues:
                raise ValueError("Ma'lumotlarni to'ldiring: " + " ; ".join(validation_issues[:10]))

            import_result = upsert_imported_students(new_students)

            request.session["uploaded"] = True
            request.session["upload_added_count"] = import_result["added"]
            request.session["upload_updated_count"] = import_result["updated"]
            request.session["upload_skipped_duplicates"] = import_result["skipped_duplicates"]
            request.session["course"] = course
            request.session["start_date"] = start_date
            request.session["end_date"] = end_date
            request.session["practice_type"] = practice_type or get_practice_type(course)
            return redirect("upload_excel")

        except Exception as exc:
            return render(
                request,
                "app_excel/upload.html",
                {"error": str(exc), "profile": profile, "template_sources": get_template_sources()},
            )

    uploaded = request.session.pop("uploaded", False)
    return render(
        request,
        "app_excel/upload.html",
        {
            "uploaded": uploaded,
            "upload_added_count": request.session.pop("upload_added_count", 0),
            "upload_updated_count": request.session.pop("upload_updated_count", 0),
            "upload_skipped_duplicates": request.session.pop("upload_skipped_duplicates", 0),
            "profile": profile,
            "template_sources": get_template_sources(),
        },
    )


@login_required(login_url="login")
def export_to_word(request):
    missing_templates = validate_runtime_templates()
    if missing_templates:
        return HttpResponse(f"Faol DOCX shablonlarda muammo bor: {', '.join(missing_templates)}.", status=400)

    students = Student.objects.all()
    if not students.exists():
        return HttpResponse("Hali hech qanday talaba ma'lumotlari mavjud emas.", status=400)

    course = int(request.GET.get("course") or request.session.get("course") or 4)
    practice_type = request.session.get("practice_type") or get_practice_type(course)
    first_student = students.first()
    start_date_str = request.session.get("start_date", "2026-02-02")
    end_date_str = request.session.get("end_date", "2026-04-11")

    validation_issues = validate_students_for_documents([first_student], course, start_date_str, end_date_str)
    if validation_issues:
        return HttpResponse(f"Ma'lumotlarni to'ldiring. Hujjat yuklanmadi: {' ; '.join(validation_issues[:10])}", status=400)

    document = generate_shartnoma_document(first_student, course, practice_type, start_date_str, end_date_str)
    return save_document_response(document, f"{sanitize_filename(first_student.company)}_shartnoma.docx")


@login_required(login_url="login")
def export_all_documents_zip(request):
    missing_templates = validate_runtime_templates()
    if missing_templates:
        return HttpResponse(f"ZIP yaratish uchun faol DOCX shablonlarda muammo bor: {', '.join(missing_templates)}.", status=400)

    students = Student.objects.all()
    if not students.exists():
        return HttpResponse("Talabalar ma'lumotlari topilmadi.", status=400)

    course = int(request.session.get("course", 4))
    start_date_str = request.session.get("start_date", "2025-02-17")
    end_date_str = request.session.get("end_date", "2025-04-26")
    validation_issues = validate_students_for_documents(students, course, start_date_str, end_date_str)
    if validation_issues:
        return HttpResponse(f"Ma'lumotlarni to'ldiring. ZIP yuklanmadi: {' ; '.join(validation_issues[:10])}", status=400)

    practice_type = request.session.get("practice_type") or get_practice_type(course)
    zip_buffer = build_documents_zip(students, course, practice_type, start_date_str, end_date_str)
    response = HttpResponse(zip_buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="Hujjatlar.zip"'
    return response


@login_required(login_url="login")
def generate_contract_for_company(request, company_name):
    missing_templates = validate_runtime_templates()
    if missing_templates:
        return HttpResponse(f"Faol DOCX shablonlarda muammo bor: {', '.join(missing_templates)}.", status=400)

    students = Student.objects.filter(company=company_name)
    if not students.exists():
        return HttpResponse("Bu korxona bo'yicha talabalar topilmadi.", status=404)

    course = int(request.GET.get("course") or request.session.get("course") or 4)
    practice_type = request.session.get("practice_type") or get_practice_type(course)
    first_student = students.first()
    start_date_str = request.session.get("start_date", "2026-02-02")
    end_date_str = request.session.get("end_date", "2026-04-11")
    document = generate_shartnoma_document(first_student, course, practice_type, start_date_str, end_date_str)
    return save_document_response(document, f"{sanitize_filename(company_name)}_shartnoma.docx")


@login_required(login_url="login")
def download_student_document(request, student_id, document_type):
    missing_templates = validate_runtime_templates()
    if missing_templates:
        return HttpResponse(f"Faol DOCX shablonlarda muammo bor: {', '.join(missing_templates)}.", status=400)

    student = get_object_or_404(Student, pk=student_id)
    course = int(request.GET.get("course") or request.session.get("course") or 4)
    practice_type = request.session.get("practice_type") or get_practice_type(course)
    start_date_str = request.session.get("start_date", "2025-02-17")
    end_date_str = request.session.get("end_date", "2025-04-26")

    validation_issues = validate_students_for_documents([student], course, start_date_str, end_date_str)
    if validation_issues:
        return HttpResponse(f"Ma'lumotlarni to'ldiring. Hujjat yuklanmadi: {' ; '.join(validation_issues[:10])}", status=400)

    generators = {
        "shartnoma": generate_shartnoma_document,
        "kundalik": generate_kundalik_document,
        "yollanma": generate_yollanma_document,
    }
    generator = generators.get(document_type)
    if generator is None:
        return HttpResponse("Hujjat turi topilmadi.", status=404)

    document = generator(student, course, practice_type, start_date_str, end_date_str)
    filename = f"{sanitize_filename(student.full_name)}_{document_type}.docx"
    return save_document_response(document, filename)


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        remember = request.POST.get("remember")
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            ensure_profile(user)
            request.session.set_expiry(31536000 if remember else 0)
            return redirect("dashboard")

        messages.error(request, "Login yoki parol noto'g'ri.")
    return render(request, "app_excel/login.html")


def register_view(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if not full_name:
            messages.error(request, "F.I.Sh. kiritilishi kerak.")
        elif not email:
            messages.error(request, "Email manzilini kiriting.")
        elif password1 != password2:
            messages.error(request, "Parollar mos emas!")
        elif User.objects.filter(username=username).exists():
            messages.error(request, "Bu login allaqachon mavjud!")
        elif User.objects.filter(email=email).exists():
            messages.error(request, "Bu email allaqachon mavjud!")
        elif len(password1 or "") < 8:
            messages.error(request, "Parol kamida 8 ta belgidan iborat bo'lishi kerak.")
        else:
            name_parts = full_name.split(maxsplit=1)
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=name_parts[0],
                last_name=name_parts[1] if len(name_parts) > 1 else "",
            )
            ensure_profile(user).save()
            messages.success(request, "Muvaffaqiyatli ro'yxatdan o'tildi.")
            login(request, user)
            return redirect("dashboard")

    return render(request, "app_excel/register.html")


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required(login_url="login")
def profile_view(request):
    profile = ensure_profile(request.user)
    return render(request, "app_excel/profile.html", {"user": request.user, "profile": profile})


@login_required(login_url="login")
def account_settings_view(request):
    profile = ensure_profile(request.user)
    if request.method == "POST":
        action = request.POST.get("action", "password")

        if action == "profile":
            username = request.POST.get("username", "").strip()
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            email = request.POST.get("email", "").strip()

            if not username:
                messages.error(request, "Foydalanuvchi nomi bo'sh bo'lishi mumkin emas.")
            elif User.objects.exclude(pk=request.user.pk).filter(username=username).exists():
                messages.error(request, "Bu foydalanuvchi nomi allaqachon band.")
            elif email and User.objects.exclude(pk=request.user.pk).filter(email__iexact=email).exists():
                messages.error(request, "Bu email allaqachon boshqa akkauntga ulangan.")
            else:
                request.user.username = username
                request.user.first_name = first_name
                request.user.last_name = last_name
                request.user.email = email
                request.user.save(update_fields=["username", "first_name", "last_name", "email"])
                messages.success(request, "Profil ma'lumotlari muvaffaqiyatli yangilandi.")
                return redirect("account_settings")
        else:
            current_password = request.POST.get("current_password", "")
            new_password = request.POST.get("new_password", "")
            confirm_password = request.POST.get("confirm_password", "")

            if not request.user.check_password(current_password):
                messages.error(request, "Joriy parol noto'g'ri.")
            elif len(new_password) < 8:
                messages.error(request, "Yangi parol kamida 8 ta belgidan iborat bo'lishi kerak.")
            elif new_password != confirm_password:
                messages.error(request, "Yangi parollar bir-biriga mos emas.")
            elif current_password == new_password:
                messages.error(request, "Yangi parol joriy paroldan farq qilishi kerak.")
            else:
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, "Parol muvaffaqiyatli yangilandi.")
                return redirect("account_settings")

    return render(request, "app_excel/account_settings.html", {"user": request.user, "profile": profile})


def handler403(request, exception=None):
    return render(request, "app_excel/errors/403.html", status=403)


def handler404(request, exception=None):
    return render(request, "app_excel/errors/404.html", status=404)


def handler500(request):
    return render(request, "app_excel/errors/500.html", status=500)


def csrf_failure(request, reason=""):
    return render(request, "app_excel/403_csrf.html", status=403)
