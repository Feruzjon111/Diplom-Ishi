import json
from io import BytesIO

import openpyxl
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse

from .models import Student


class DefaultChatViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="chatuser", password="StrongPass123")
        self.url = reverse("ai_chat")

    def test_chat_requires_login(self):
        response = self.client.post(self.url, data="{}", content_type="application/json")

        self.assertEqual(response.status_code, 302)

    def test_chat_rejects_empty_message(self):
        self.client.force_login(self.user)

        response = self.client.post(self.url, data="{}", content_type="application/json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("error", response.json())

    def test_chat_returns_default_excel_help(self):
        self.client.force_login(self.user)

        response = self.client.post(
            self.url,
            data=json.dumps({"message": "Excel qanday yuklanadi?"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("/excel/upload/", response.json()["reply"])

    def test_chat_returns_error_help_for_error_words(self):
        self.client.force_login(self.user)

        response = self.client.post(
            self.url,
            data=json.dumps({"message": "xato chiqdi"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("Agar xatolik chiqsa", response.json()["reply"])

    @override_settings(SUPPORT_TELEGRAM="@feruzjon", SUPPORT_PHONE="956561212")
    def test_chat_fallback_includes_support_contacts(self):
        self.client.force_login(self.user)

        response = self.client.post(
            self.url,
            data=json.dumps({"message": "Buni tushunmadim"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("@feruzjon", response.json()["reply"])
        self.assertIn("https://t.me/feruzjon", response.json()["reply"])
        self.assertIn("+998 95 656 12 12", response.json()["reply"])


class UploadExcelViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="uploader", password="StrongPass123")
        self.url = reverse("upload_excel")

    def test_sample_excel_contains_ten_students_with_mixed_courses_and_companies(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("download_sample_excel"))

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook["Talabalar"]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        courses = {row[4] for row in data_rows}
        companies = {row[8] for row in data_rows}

        self.assertEqual(len(data_rows), 10)
        self.assertEqual(courses, {3, 4})
        self.assertEqual(len(companies), 10)

    def build_excel_file(self, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "Talabaning F.I.Sh.",
                "Fakultet",
                "Yo'nalish",
                "Guruh",
                "Kurs",
                "Amaliyot turi",
                "Boshlanish sanasi",
                "Tugash sanasi",
                "Korxona nomi",
                "Korxona manzili",
                "Korxona rahbari F.I.Sh.",
                "Korxona telefoni",
                "Universitet amaliyot rahbari",
                "Fakultet dekani",
                "Kafedra mudiri",
            ]
        )
        for row in rows:
            sheet.append(row)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            "talabalar.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_upload_appends_students_without_deleting_existing_ones(self):
        Student.objects.create(
            full_name="Aliyev Bekzod Anvar o'g'li",
            faculty="Axborot texnologiyalari",
            direction="Dasturiy injiniring",
            group="SE-401",
            company="TechSoft MCHJ",
            company_address="Eski manzil",
            company_director="Eski rahbar",
            company_phone="+998900000000",
            practice_supervisor="Rasulov Dilshod Qodirovich",
            faculty_dean="O.B. Ro'zibayev",
            department_head="N.O. Raximov",
        )
        self.client.force_login(self.user)

        upload_file = self.build_excel_file(
            [
                [
                    "Aliyev Bekzod Anvar o'g'li",
                    "Axborot texnologiyalari",
                    "Dasturiy injiniring",
                    "SE-401",
                    4,
                    "Bitiruv oldi amaliyoti",
                    "2025-02-17",
                    "2025-04-26",
                    "TechSoft MCHJ",
                    "Yangi manzil",
                    "Yangi rahbar",
                    "+998901112233",
                    "Rasulov Dilshod Qodirovich",
                    "O.B. Ro'zibayev",
                    "N.O. Raximov",
                ],
                [
                    "Karimova Maftuna Jamshid qizi",
                    "Axborot texnologiyalari",
                    "Kompyuter injiniring",
                    "KI-402",
                    4,
                    "Bitiruv oldi amaliyoti",
                    "2025-02-17",
                    "2025-04-26",
                    "Digital Systems",
                    "Toshkent sh., Chilonzor tumani",
                    "Ergashev Oybek Bahodirovich",
                    "+998901112234",
                    "Rasulov Dilshod Qodirovich",
                    "O.B. Ro'zibayev",
                    "N.O. Raximov",
                ],
            ]
        )

        response = self.client.post(self.url, {"file": upload_file})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Student.objects.count(), 2)
        self.assertTrue(Student.objects.filter(full_name="Karimova Maftuna Jamshid qizi").exists())
        updated_student = Student.objects.get(full_name="Aliyev Bekzod Anvar o'g'li")
        self.assertEqual(updated_student.company_address, "Yangi manzil")
        self.assertEqual(updated_student.company_director, "Yangi rahbar")
